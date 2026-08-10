"""
Report Service - Report generation and management
"""
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import uuid
import json
from pathlib import Path
import io

from app.config import settings
from app.core.database_service import database_service
from app.core.websocket import ws_manager
from app.services.analysis_service import analysis_service


class ReportService:
    """
    Service for report operations
    """
    
    def __init__(self):
        self.reports_dir = Path(settings.UPLOAD_DIR) / "reports"
        self.reports_dir.mkdir(parents=True, exist_ok=True)
    
    async def generate_report(
        self,
        user_id: str,
        upload_id: str,
        report_type: str,
        report_format: str,
        filters: Optional[Dict] = None
    ) -> Dict:
        """
        Generate a new report
        """
        # Verify upload ownership
        upload = await database_service.get_upload_by_id(upload_id)
        if not upload or upload.get("user_id") != user_id:
            raise ValueError("Upload not found")
        
        report_id = str(uuid.uuid4())
        report_name = f"{report_type}_{upload.get('name', 'report')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Create report record
        report = await database_service.create_report({
            "id": report_id,
            "user_id": user_id,
            "upload_id": upload_id,
            "name": report_name,
            "type": report_type,
            "format": report_format,
            "status": "generating",
            "filters": filters,
            "created_at": datetime.utcnow().isoformat()
        })
        
        # Generate report asynchronously (simplified - use background task in production)
        try:
            await self._generate_report_content(
                report_id, user_id, upload_id, report_type, report_format, filters
            )
        except Exception as e:
            await database_service.update_report(report_id, {
                "status": "failed",
                "error": str(e)
            })
            await ws_manager.broadcast_report_status(
                user_id, report_id, "failed"
            )
        
        return {
            "reportId": report_id,
            "status": "generating",
            "estimatedTime": 30  # seconds
        }
    
    async def _generate_report_content(
        self,
        report_id: str,
        user_id: str,
        upload_id: str,
        report_type: str,
        report_format: str,
        filters: Optional[Dict]
    ):
        """
        Generate the actual report content
        """
        # Get analysis data
        analysis = await analysis_service.get_analysis_results(upload_id, user_id)
        
        if not analysis:
            raise ValueError("No analysis data found")
        
        # Generate based on format
        if report_format == "json":
            content = await self._generate_json_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.json"
            with open(file_path, 'w') as f:
                json.dump(content, f, indent=2, default=str)
        
        elif report_format == "excel":
            content = await self._generate_excel_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.xlsx"
            content.save(file_path)
        
        elif report_format == "pdf":
            content = await self._generate_pdf_report(analysis, report_type, filters)
            file_path = self.reports_dir / f"{report_id}.pdf"
            with open(file_path, 'wb') as f:
                f.write(content)
        
        # Update report record
        file_size = file_path.stat().st_size if file_path.exists() else 0
        
        await database_service.update_report(report_id, {
            "status": "completed",
            "file_path": str(file_path),
            "size": file_size,
            "completed_at": datetime.utcnow().isoformat()
        })
        
        # Notify via WebSocket
        await ws_manager.broadcast_report_status(
            user_id, report_id, "completed",
            f"/api/v1/reports/download/{report_id}"
        )
    
    async def _generate_json_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ) -> Dict:
        """
        Generate JSON report
        """
        report = {
            "generatedAt": datetime.utcnow().isoformat(),
            "reportType": report_type,
            "summary": analysis.get("summary", {}),
        }
        
        if report_type in ["compliance", "investigation"]:
            report["patterns"] = analysis.get("patterns", [])
            report["suspiciousAddresses"] = analysis.get("suspiciousAddresses", [])
        
        if report_type == "compliance":
            report["complianceNotes"] = self._generate_compliance_notes(analysis)
        
        return report
    
    async def _generate_excel_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ):
        """
        Generate Excel report
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary = analysis.get("summary", {})
        ws_summary.append(["Metric", "Value"])
        for key, value in summary.items():
            ws_summary.append([key, str(value)])
        
        # Patterns sheet
        if analysis.get("patterns"):
            ws_patterns = wb.create_sheet("Patterns")
            patterns_df = pd.DataFrame(analysis["patterns"])
            
            for r_idx, row in enumerate(dataframe_to_rows(patterns_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_patterns.cell(row=r_idx, column=c_idx, value=str(value) if isinstance(value, list) else value)
        
        # Suspicious Addresses sheet
        if analysis.get("suspiciousAddresses"):
            ws_addresses = wb.create_sheet("Suspicious Addresses")
            addresses_df = pd.DataFrame(analysis["suspiciousAddresses"])
            
            for r_idx, row in enumerate(dataframe_to_rows(addresses_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_addresses.cell(row=r_idx, column=c_idx, value=str(value) if isinstance(value, list) else value)
        
        return wb
    
    async def _generate_pdf_report(
        self,
        analysis: Dict,
        report_type: str,
        filters: Optional[Dict]
    ) -> bytes:
        """
        Generate a professional SAR-grade PDF report using reportlab.
        Includes FATF Red Flag mapping, compliance sections, and SmurfPakad branding.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
                HRFlowable, KeepTogether
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=letter,
                topMargin=0.7*inch, bottomMargin=0.7*inch,
                leftMargin=0.75*inch, rightMargin=0.75*inch,
            )
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'SARTitle', parent=styles['Heading1'],
                fontSize=18, spaceAfter=6, textColor=colors.HexColor('#1a0b2e'),
            )
            subtitle_style = ParagraphStyle(
                'SARSubtitle', parent=styles['Normal'],
                fontSize=10, textColor=colors.HexColor('#666666'),
            )
            section_style = ParagraphStyle(
                'SARSection', parent=styles['Heading2'],
                fontSize=13, textColor=colors.HexColor('#7c3aed'),
                spaceBefore=16, spaceAfter=8,
            )
            warning_style = ParagraphStyle(
                'SARWarning', parent=styles['Normal'],
                fontSize=10, textColor=colors.HexColor('#dc2626'),
                backColor=colors.HexColor('#fef2f2'),
                borderPadding=8, spaceAfter=10,
            )
            
            # ---- Header ----
            elements.append(Paragraph("SUSPICIOUS ACTIVITY REPORT", title_style))
            elements.append(Paragraph(
                f"SmurfPakad AML Intelligence Platform — Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
                subtitle_style
            ))
            elements.append(Paragraph(
                f"Report Type: {report_type.upper()} | Powered by IBM watsonx.ai + GNN Analysis",
                subtitle_style
            ))
            elements.append(HRFlowable(
                width="100%", thickness=2,
                color=colors.HexColor('#7c3aed'), spaceAfter=12
            ))
            
            # ---- Classification Banner ----
            summary = analysis.get("summary", {})
            suspicious_count = summary.get("suspiciousTransactions", summary.get("suspicious_count", 0))
            total_txns = summary.get("totalTransactions", summary.get("total_transactions", 0))
            
            if suspicious_count > 0:
                elements.append(Paragraph(
                    f"⚠ ALERT: {suspicious_count} suspicious transactions detected out of {total_txns} analyzed. "
                    f"SAR filing is recommended within 24 hours.",
                    warning_style
                ))
            
            # ---- Executive Summary ----
            elements.append(Paragraph("1. Executive Summary", section_style))
            summary_data = [
                ["Metric", "Value"],
                ["Total Transactions Analyzed", str(total_txns)],
                ["Suspicious Transactions", str(suspicious_count)],
                ["Critical Patterns", str(len([p for p in analysis.get("patterns", []) if p.get("severity") == "critical"]))],
                ["High-Risk Addresses", str(len([a for a in analysis.get("suspiciousAddresses", []) if a.get("riskLevel") in ("critical", "high")]))],
                ["Analysis Date", datetime.utcnow().strftime('%Y-%m-%d')],
                ["Model", "SmurfPakad GNN (GraphSAGE/GATv2)"],
            ]
            t = Table(summary_data, colWidths=[250, 200])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f3ff'), colors.white]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 16))
            
            # ---- Detected Patterns ----
            patterns = analysis.get("patterns", [])
            if patterns:
                elements.append(Paragraph("2. Detected Structural Patterns", section_style))
                
                pattern_data = [["Pattern Type", "Severity", "Description"]]
                for p in patterns[:15]:
                    pattern_data.append([
                        p.get("type", "unknown").replace("_", " ").title(),
                        p.get("severity", "medium").upper(),
                        p.get("description", "N/A")[:80],
                    ])
                
                t = Table(pattern_data, colWidths=[120, 80, 260])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f3ff'), colors.white]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 16))
            
            # ---- FATF Red Flag Mapping ----
            elements.append(Paragraph("3. FATF Red Flag Indicator Mapping", section_style))
            
            fatf_flags = self._map_patterns_to_fatf(patterns)
            if fatf_flags:
                for flag in fatf_flags:
                    elements.append(Paragraph(
                        f"<b>{flag['id']}</b> — {flag['title']}: {flag['description']}",
                        ParagraphStyle('FATFItem', parent=styles['Normal'], fontSize=9,
                                      textColor=colors.HexColor('#b45309'),
                                      spaceBefore=4, spaceAfter=4,
                                      leftIndent=12, bulletIndent=0)
                    ))
            else:
                elements.append(Paragraph(
                    "No specific FATF Red Flag Indicators matched. Alert is based on aggregate model scoring.",
                    styles['Normal']
                ))
            elements.append(Spacer(1, 16))
            
            # ---- Top Suspicious Addresses ----
            sus_addresses = analysis.get("suspiciousAddresses", [])
            if sus_addresses:
                elements.append(Paragraph("4. Top Suspicious Addresses", section_style))
                
                addr_data = [["Address", "Risk Score", "Risk Level", "In-Degree", "Out-Degree"]]
                for addr in sus_addresses[:10]:
                    addr_data.append([
                        addr.get("address", "N/A")[:20] + "...",
                        f"{addr.get('riskScore', 0):.2f}",
                        addr.get("riskLevel", "N/A").upper(),
                        str(addr.get("inDegree", 0)),
                        str(addr.get("outDegree", 0)),
                    ])
                
                t = Table(addr_data, colWidths=[120, 70, 70, 70, 70])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fef2f2'), colors.white]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 16))
            
            # ---- Compliance Notes ----
            compliance_notes = self._generate_compliance_notes(analysis)
            if compliance_notes:
                elements.append(Paragraph("5. Compliance Recommendations", section_style))
                for note in compliance_notes:
                    elements.append(Paragraph(f"• {note}", styles['Normal']))
                elements.append(Spacer(1, 12))
            
            # ---- Footer / Disclaimer ----
            elements.append(HRFlowable(
                width="100%", thickness=1,
                color=colors.HexColor('#d1d5db'), spaceAfter=8, spaceBefore=16
            ))
            footer_style = ParagraphStyle(
                'Footer', parent=styles['Normal'],
                fontSize=7, textColor=colors.HexColor('#9ca3af'),
                alignment=TA_CENTER,
            )
            elements.append(Paragraph(
                f"This report was generated by SmurfPakad v{settings.APP_VERSION} "
                f"using IBM watsonx.ai and GNN-based analysis. "
                f"This document is confidential and intended for authorized compliance personnel only. "
                f"Report ID: {uuid.uuid4().hex[:12]} | Generated: {datetime.utcnow().isoformat()}",
                footer_style
            ))
            
            doc.build(elements)
            return buffer.getvalue()
            
        except ImportError:
            return b"PDF generation requires reportlab. Install with: pip install reportlab"
    
    def _map_patterns_to_fatf(self, patterns: List[Dict]) -> List[Dict]:
        """Map patterns to FATF indicators for the SAR report."""
        try:
            from app.services.fatf_service import fatf_service
            return fatf_service.map_patterns_to_fatf(patterns)
        except Exception:
            # Fallback inline mapping
            flags = []
            types = {p.get("type", "") for p in patterns}
            if "fan_out" in types:
                flags.append({"id": "FATF-3.1", "title": "Structuring", "description": "Transaction splitting detected."})
            if "fan_in" in types:
                flags.append({"id": "FATF-3.2", "title": "Aggregation", "description": "Fund consolidation detected."})
            if "pass_through" in types:
                flags.append({"id": "FATF-5.1", "title": "Mule Activity", "description": "Pass-through wallet detected."})
            return flags
    
    def _generate_compliance_notes(self, analysis: Dict) -> List[str]:
        """
        Generate compliance notes based on analysis
        """
        notes = []
        summary = analysis.get("summary", {})
        
        if summary.get("suspiciousTransactions", 0) > 0:
            notes.append(
                f"ALERT: {summary['suspiciousTransactions']} suspicious transactions detected. "
                "Manual review recommended."
            )
        
        patterns = analysis.get("patterns", [])
        critical_patterns = [p for p in patterns if p.get("severity") == "critical"]
        if critical_patterns:
            notes.append(
                f"CRITICAL: {len(critical_patterns)} critical patterns detected. "
                "Immediate investigation required."
            )
        
        high_risk_addresses = [
            a for a in analysis.get("suspiciousAddresses", [])
            if a.get("riskLevel") in ["critical", "high"]
        ]
        if high_risk_addresses:
            notes.append(
                f"HIGH RISK: {len(high_risk_addresses)} addresses flagged as high/critical risk. "
                "Consider filing SAR."
            )
        
        return notes
    
    async def get_reports(
        self,
        user_id: str,
        upload_id: Optional[str] = None,
        page: int = 1,
        limit: int = 10
    ) -> Tuple[List[Dict], int]:
        """Get reports for a user"""
        reports, total = await database_service.get_reports_by_user(
            user_id=user_id,
            upload_id=upload_id,
            page=page,
            limit=limit
        )
        return reports, total or 0
    
    async def get_report_by_id(
        self,
        report_id: str,
        user_id: str
    ) -> Optional[Dict]:
        """Get report by ID"""
        report = await database_service.get_report_by_id(report_id, user_id)
        return report
    
    async def get_report_file(
        self,
        report_id: str,
        user_id: str
    ) -> Optional[Dict]:
        """Get report file for download"""
        report = await database_service.get_report_by_id(report_id, user_id)
        
        if not report or report.get("status") != "completed":
            return None
        
        file_path = Path(report.get("file_path", ""))
        if not file_path.exists():
            return None
        
        with open(file_path, 'rb') as f:
            content = f.read()
        
        format_type = report.get("format", "json")
        content_types = {
            "pdf": "application/pdf",
            "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "json": "application/json"
        }
        extensions = {"pdf": ".pdf", "excel": ".xlsx", "json": ".json"}
        
        return {
            "content": content,
            "filename": f"{report['name']}{extensions.get(format_type, '')}",
            "media_type": content_types.get(format_type, "application/octet-stream")
        }
    
    async def delete_report(
        self,
        report_id: str,
        user_id: str
    ) -> bool:
        """Delete a report"""
        report = await database_service.get_report_by_id(report_id, user_id)
        
        if not report:
            return False
        
        file_path = Path(report.get("file_path", ""))
        if file_path.exists():
            file_path.unlink()
        
        await database_service.delete_report(report_id, user_id)
        return True


# Global service instance
report_service = ReportService()
